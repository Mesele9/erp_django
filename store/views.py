from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import F, Sum, Q, Avg
import openpyxl
from openpyxl.utils import get_column_letter
from .models import Item, PurchaseRecord, PurchaseRecordItem, IssueRecord, IssueRecordItem, Subcategory, Supplier
from .forms import IssueRecordFilterForm, ItemForm, ItemFilterForm, ItemsIssuedReportForm, ItemsPurchasedReportForm, PurchaseRecordFilterForm, PurchaseRecordForm, PurchaseRecordItemFormSet, IssueRecordForm, IssueRecordItemFormSet, SummarizedItemsIssuedReportForm, SummarizedItemsPurchasedReportForm, SupplierForm, SupplierFilterForm
from .resource import ItemResource
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Prefetch
from common_user.decorators import role_required


@role_required('store_staff')
@login_required
def search_items(request):
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        q = request.GET.get('q', '')
        items = Item.objects.filter(Q(description__icontains=q) | Q(category__name__icontains=q) | Q(subcategory__name__icontains=q))
        results = [{'id': item.id, 'text': item.description} for item in items]
        return JsonResponse({'results': results}, safe=False)
    return JsonResponse({'results': []}, safe=False)


@role_required('store_staff')
@login_required
def store_dashboard(request):
    #items = Item.objects.all()
    recent_purchase_records = PurchaseRecord.objects.order_by('-date')[:5]
    recent_issue_records = IssueRecord.objects.order_by('-date')[:5]
    low_stock_items = Item.objects.filter(stock_balance__lt=F('minimum_stock'))
    
    context = {
        #'items': items,
        'recent_purchase_records': recent_purchase_records,
        'recent_issue_records': recent_issue_records,
        'low_stock_items': low_stock_items,
    }
    
    return render(request, 'store/dashboard.html', context)


def load_subcategories(request):
    category_id = request.GET.get('category_id')
    if category_id:
        subcategories = Subcategory.objects.filter(category_id=category_id).order_by('name')
    else:
        subcategories = Subcategory.objects.none()
    return JsonResponse(list(subcategories.values('id', 'name')), safe=False)


@role_required('store_staff')
@login_required
def supplier_list(request):
    form = SupplierFilterForm(request.GET or None)
    suppliers = Supplier.objects.all().order_by('id')

    if form.is_valid():
        if form.cleaned_data['name']:
            suppliers = suppliers.filter(name__icontains=form.cleaned_data['name'])
        if form.cleaned_data['tin_number']:
            suppliers = suppliers.filter(tin_number__icontains=form.cleaned_data['tin_number'])

    
    paginator = Paginator(suppliers, 10)  # Show 10 suppliers per page
    page = request.GET.get('page')
    try:
        suppliers = paginator.page(page)
    except PageNotAnInteger:
        suppliers = paginator.page(1)
    except EmptyPage:
        suppliers = paginator.page(paginator.num_pages)

    return render(request, 'store/supplier_list.html', {'suppliers': suppliers, 'form': form})


@role_required('store_staff')
@login_required
def supplier_create(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('supplier_list')
    else:
        form = SupplierForm()
    return render(request, 'store/supplier_form.html', {'form': form})


@role_required('store_staff')
@login_required
def supplier_edit(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            return redirect('supplier_list')
    else:
        form = SupplierForm(instance=supplier)
    return render(request, 'store/supplier_form.html', {'form': form})


@role_required('store_staff')
@login_required
def supplier_delete(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        supplier.delete()
        return redirect('supplier_list')
    return render(request, 'store/supplier_confirm_delete.html', {'supplier': supplier})


@role_required('store_staff')
@login_required
def item_list(request):
    form = ItemFilterForm(request.GET or None)
    items = Item.objects.select_related('category', 'subcategory').only(
        'id', 'description', 'category__name', 'subcategory__name', 'unit_of_measurement', 
        'current_unit_price', 'stock_balance', 'minimum_stock'
    ).order_by('id')

    if form.is_valid() and request.method == 'GET':
        if form.cleaned_data.get('description'):
            items = items.filter(description__icontains=form.cleaned_data['description'])
        if form.cleaned_data.get('category'):
            items = items.filter(category=form.cleaned_data['category'])
        if form.cleaned_data.get('subcategory'):
            items = items.filter(subcategory=form.cleaned_data['subcategory'])


    paginator = Paginator(items, 25)  # Show 25 items per page
    page = request.GET.get('page')
    try:
        items = paginator.page(page)
    except PageNotAnInteger:
        items = paginator.page(1)
    except EmptyPage:
        items = paginator.page(paginator.num_pages)

    context = {
        'items': items,
        'form': form,
        'selected_category': request.GET.get('category'),
        'selected_subcategory': request.GET.get('subcategory'),
    }

    return render(request, 'store/item_list.html', context)


@role_required('store_staff')
@login_required
def item_create(request):
    if request.method == 'POST':
        form = ItemForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('item_list')
    else:
        form = ItemForm()
    return render(request, 'store/item_form.html', {'form': form})


@role_required('store_staff')
@login_required
def item_edit(request, pk):
    item = get_object_or_404(Item, pk=pk)
    if request.method == 'POST':
        form = ItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return redirect('item_list')
    else:
        form = ItemForm(instance=item)
    return render(request, 'store/item_form.html', {'form': form})


@role_required('store_staff')
@login_required
def item_delete(request, pk):
    item = Item.objects.get(pk=pk)
    if request.method == 'POST':
        item.delete()
        return redirect('item_list')
    return render(request, 'store/item_confirm_delete.html', {'item': item})


@role_required('store_staff')
@login_required
def purchase_record_list(request):
    form = PurchaseRecordFilterForm(request.GET or None)
    purchase_records = PurchaseRecord.objects.select_related('supplier').prefetch_related('items').order_by('-date')

    if form.is_valid():
        if form.cleaned_data['date_from']:
            purchase_records = purchase_records.filter(date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data['date_to']:
            purchase_records = purchase_records.filter(date__lte=form.cleaned_data['date_to'])
        if form.cleaned_data['voucher_number']:
            purchase_records = purchase_records.filter(voucher_number__icontains=form.cleaned_data['voucher_number'])

    paginator = Paginator(purchase_records, 20)  # Show 20 records per page
    page = request.GET.get('page')
    try:
        purchase_records = paginator.page(page)
    except PageNotAnInteger:
        purchase_records = paginator.page(1)
    except EmptyPage:
        purchase_records = paginator.page(paginator.num_pages)
    

    return render(request, 'store/purchase_record_list.html', {'purchase_records': purchase_records, 'form': form})


@role_required('store_staff')
@login_required
def purchase_record_detail(request, pk):
    purchase_record = get_object_or_404(PurchaseRecord.objects.prefetch_related('items'), pk=pk)
    return render(request, 'store/purchase_record_detail.html', {'purchase_record': purchase_record})


@role_required('store_staff')
@login_required
def purchase_record_create(request):
    if request.method == 'POST':
        form = PurchaseRecordForm(request.POST, request.FILES)
        formset = PurchaseRecordItemFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            purchase_record = form.save()
            items = formset.save(commit=False)
            for item in items:
                item.purchase_record = purchase_record
                item.save()
            return redirect('purchase_record_list')
    else:
        form = PurchaseRecordForm()
        formset = PurchaseRecordItemFormSet()

    return render(request, 'store/purchase_record_form.html', {'form': form, 'formset': formset})


@role_required('store_staff')
@login_required
def purchase_record_edit(request, pk):
    purchase_record = get_object_or_404(PurchaseRecord, pk=pk)
    if request.method == 'POST':
        form = PurchaseRecordForm(request.POST, request.FILES, instance=purchase_record)
        formset = PurchaseRecordItemFormSet(request.POST, instance=purchase_record)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            return redirect('purchase_record_list')
    else:
        form = PurchaseRecordForm(instance=purchase_record)
        formset = PurchaseRecordItemFormSet(instance=purchase_record)
    return render(request, 'store/purchase_record_form.html', {'form': form, 'formset': formset})


@role_required('store_staff')
@login_required
def purchase_record_delete(request, pk):
    purchase_record = get_object_or_404(PurchaseRecord, pk=pk)
    if request.method == 'POST':
        purchase_record.delete()
        return redirect('purchase_record_list')
    return render(request, 'store/purchase_record_confirm_delete.html', {'purchase_record': purchase_record})


@role_required('store_staff')
@login_required
def issue_record_list(request):
    form = IssueRecordFilterForm(request.GET or None)
    
    issue_records = IssueRecord.objects.select_related('department').prefetch_related('items').order_by('-date')

    if form.is_valid():
        if form.cleaned_data['date_from']:
            issue_records = issue_records.filter(date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data['date_to']:
            issue_records = issue_records.filter(date__lte=form.cleaned_data['date_to'])
        if form.cleaned_data['voucher_number']:
            issue_records = issue_records.filter(voucher_number__icontains=form.cleaned_data['voucher_number'])

    paginator = Paginator(issue_records, 20)  # Show 20 records per page
    page = request.GET.get('page')
    try:
        issue_records = paginator.page(page)
    except PageNotAnInteger:
        issue_records = paginator.page(1)
    except EmptyPage:
        issue_records = paginator.page(paginator.num_pages)

    return render(request, 'store/issue_record_list.html', {'issue_records': issue_records, 'form': form})


@role_required('store_staff')
@login_required
def issue_record_detail(request, pk):
    issue_record = get_object_or_404(IssueRecord, pk=pk)
    issue_record_items = IssueRecordItem.objects.filter(issue_record=issue_record)
    return render(request, 'store/issue_record_detail.html', {'issue_record': issue_record, 'issue_record_items': issue_record_items})


@role_required('store_staff')
@login_required
def issue_record_create(request):
    if request.method == 'POST':
        form = IssueRecordForm(request.POST)
        formset = IssueRecordItemFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            issue_record = form.save()
            items = formset.save(commit=False)
            for item in items:
                item.issue_record = issue_record
                item.save()
            return redirect('issue_record_list')
    else:
        form = IssueRecordForm()
        formset = IssueRecordItemFormSet()

    return render(request, 'store/issue_record_form.html', {'form': form, 'formset': formset})


@role_required('store_staff')
@login_required
def issue_record_edit(request, pk):
    issue_record = get_object_or_404(IssueRecord, pk=pk)
    if request.method == 'POST':
        form = IssueRecordForm(request.POST, instance=issue_record)
        formset = IssueRecordItemFormSet(request.POST, instance=issue_record)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            return redirect('issue_record_list')
    else:
        form = IssueRecordForm(instance=issue_record)
        formset = IssueRecordItemFormSet(instance=issue_record)
    return render(request, 'store/issue_record_form.html', {'form': form, 'formset': formset})


@role_required('store_staff')
@login_required
def issue_record_delete(request, pk):
    issue_record = get_object_or_404(IssueRecord, pk=pk)
    if request.method == 'POST':
        issue_record.delete()
        return redirect('issue_record_list')
    return render(request, 'store/issue_record_confirm_delete.html', {'issue_record': issue_record})


@role_required('store_staff')
@login_required
def items_purchased_report(request):
    form = ItemsPurchasedReportForm(request.GET or None)
    report_data = []

    if form.is_valid():
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        voucher_number = form.cleaned_data.get('voucher_number')
        supplier = form.cleaned_data.get('supplier')
        description = form.cleaned_data.get('description')

        filters = Q()
        if date_from:
            filters &= Q(purchase_record__date__gte=date_from)
        if date_to:
            filters &= Q(purchase_record__date__lte=date_to)
        if voucher_number:
            filters &= Q(purchase_record__voucher_number__icontains=voucher_number)
        if supplier:
            filters &= Q(purchase_record__supplier=supplier)
        if description:
            filters &= Q(item__description__icontains=description)

        report_data = PurchaseRecordItem.objects.filter(filters).values(
            'item__description',
            'item__category__name',
            'item__subcategory__name',
            'purchase_record__date',
            'purchase_record__voucher_number',
            'purchase_record__supplier__name',
        ).annotate(
            total_quantity=Sum('quantity'),
            total_price=Sum('total_price')
        )

        if 'export' in request.GET and request.GET['export'] == 'excel':
            return export_to_excel(report_data, 'Items Purchased Report')

    return render(request, 'store/items_purchased_report.html', {
        'form': form,
        'report_data': report_data,
    })


@role_required('store_staff')
@login_required
def items_issued_report(request):
    form = ItemsIssuedReportForm(request.GET or None)
    report_data = []

    if form.is_valid():
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        voucher_number = form.cleaned_data.get('voucher_number')
        department = form.cleaned_data.get('department')
        description = form.cleaned_data.get('description')

        filters = Q()
        if date_from:
            filters &= Q(issue_record__date__gte=date_from)
        if date_to:
            filters &= Q(issue_record__date__lte=date_to)
        if voucher_number:
            filters &= Q(issue_record__voucher_number__icontains=voucher_number)
        if department:
            filters &= Q(issue_record__department=department)
        if description:
            filters &= Q(item__description__icontains=description)

        report_data = IssueRecordItem.objects.filter(filters).values(
            'item__description',
            'item__category__name',
            'item__subcategory__name',
            'issue_record__date',
            'issue_record__voucher_number',
            'issue_record__department__name',
        ).annotate(
            total_quantity=Sum('quantity'),
            total_price=Sum(F('quantity') * F('item__current_unit_price'))

        )

        if 'export' in request.GET and request.GET['export'] == 'excel':
            return export_to_excel(report_data, 'Items Issued Report')

    return render(request, 'store/items_issued_report.html', {
        'form': form,
        'report_data': report_data,
    })


@role_required('store_staff')
@login_required
def summarized_items_purchased_report(request):
    form = SummarizedItemsPurchasedReportForm(request.GET or None)
    report_data = []

    if form.is_valid():
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        supplier = form.cleaned_data.get('supplier')

        filters = Q()
        if date_from:
            filters &= Q(purchase_record__date__gte=date_from)
        if date_to:
            filters &= Q(purchase_record__date__lte=date_to)
        if supplier:
            filters &= Q(purchase_record__supplier=supplier)

        report_data = PurchaseRecordItem.objects.filter(filters).values(
            'item__description',
            'item__category__name',
            'item__subcategory__name'
        ).annotate(
            total_quantity=Sum('quantity'),
            total_price=Sum('total_price')
        )

    if 'export' in request.GET:
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=summarized_items_purchased_report.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Summarized Items Purchased'

        headers = ['Item Description', 'Category', 'Subcategory', 'Total Quantity', 'Total Price']
        ws.append(headers)

        for record in report_data:
            row = [
                record['item__description'],
                record['item__category__name'],
                record['item__subcategory__name'],
                record['total_quantity'],
                record['total_price'],
            ]
            ws.append(row)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].auto_size = True

        wb.save(response)
        return response

    return render(request, 'store/summarized_items_purchased_report.html', {
        'form': form,
        'report_data': report_data,
    })

from django.db.models import F, FloatField, ExpressionWrapper


@role_required('store_staff')
@login_required
def summarized_items_issued_report(request):
    form = SummarizedItemsIssuedReportForm(request.GET or None)
    report_data = []

    if form.is_valid():
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        department = form.cleaned_data.get('department')

        filters = Q()
        if date_from:
            filters &= Q(issue_record__date__gte=date_from)
        if date_to:
            filters &= Q(issue_record__date__lte=date_to)
        if department:
            filters &= Q(issue_record__department=department)

        report_data = IssueRecordItem.objects.filter(filters).values(
            'item__description',
            'item__category__name',
            'item__subcategory__name'
        ).annotate(
            total_quantity=Sum('quantity'),
            total_price=Sum(F('quantity') * F('item__current_unit_price'))
        )

    if 'export' in request.GET:
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=summarized_items_issued_report.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Summarized Items Issued'

        headers = ['Item Description', 'Category', 'Subcategory', 'Total Quantity', 'Total Price']
        ws.append(headers)

        for record in report_data:
            row = [
                record['item__description'],
                record['item__category__name'],
                record['item__subcategory__name'],
                record['total_quantity'],
                record['total_price'],
                
            ]
            ws.append(row)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].auto_size = True

        wb.save(response)
        return response

    return render(request, 'store/summarized_items_issued_report.html', {
        'form': form,
        'report_data': report_data,
    })


@role_required('store_staff')
@login_required
def export_to_excel(data, report_name):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = report_name

    headers = list(data[0].keys()) if data else []
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        worksheet[f'{column_letter}1'] = header

    for row_num, item in enumerate(data, 2):
        for col_num, (key, value) in enumerate(item.items(), 1):
            column_letter = get_column_letter(col_num)
            worksheet[f'{column_letter}{row_num}'] = value

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={report_name}.xlsx'
    workbook.save(response)
    return response


@role_required('store_staff')
@login_required
def export_items_to_excel(request):
    item_resource = ItemResource()
    dataset = item_resource.export()
    response = HttpResponse(dataset.xlsx, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="items.xlsx"'
    return response
